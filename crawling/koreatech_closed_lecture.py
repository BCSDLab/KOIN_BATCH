import pymysql
import urllib3
import openpyxl
import config
import time

### static field ###
# 폐강된 강좌 엑셀파일
filename = 'lecture_closed.xlsx'  # 읽어들일 엑셀파일명
start_row = 4  # 데이터가 시작하는 row
year_col = 'B'  # 학년도 column
semester_col = 'C'  # 학기 column
code_col = 'D'  # 교과목코드 column
name_col = 'E'  # 교과목명 column
grades_col = 'G'  # 학점 column
class_number_col = 'K'  # 분반 column
department_col = 'N'  # 개설학과 column
professor_col = 'M'  # 교수 column
is_english_col = 'O'  # 영어강의여부 column


def connect_db():
    urllib3.disable_warnings()
    conn = pymysql.connect(host=config.DATABASE_CONFIG['host'],
                           port=config.DATABASE_CONFIG['port'],
                           user=config.DATABASE_CONFIG['user'],
                           password=config.DATABASE_CONFIG['password'],
                           db=config.DATABASE_CONFIG['db'],
                           charset='utf8')
    return conn


def crawling():
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.active
    lectures = []
    year = ws['%s%d' % (year_col, start_row)].value
    semester = ws['%s%d' % (semester_col, start_row)].value
    semester_date = '%s%s' % (year, semester.split('학기')[0])

    for row in range(start_row, ws.max_row + 1):
        code = ws['%s%d' % (code_col, row)].value
        name = ws['%s%d' % (name_col, row)].value
        grades = ws['%s%d' % (grades_col, row)].value
        class_number = ws['%s%d' % (class_number_col, row)].value
        department = ws['%s%d' % (department_col, row)].value
        professor = ws['%s%d' % (professor_col, row)].value
        is_english = ws['%s%d' % (is_english_col, row)].value
        lecture = Lecture(semester_date=semester_date, code=code, name=name, grades=grades, class_number=class_number,
                          department=department, professor=professor, is_english=is_english)
        lectures.append(lecture)

        # print(semester_date, code, name, grades, class_number, department, professor, is_english)

    updateDB(lectures, semester_date)
    pass


def updateDB(lectures, semester_date):
    cur = connection.cursor()
    try:
        for lecture in lectures:
                sql = "DELETE FROM koin.lectures WHERE semester_date='%s' and code='%s' and name='%s' and grades='%s' and class='%s' and department='%s' and professor='%s' and is_english='%s'"

                cur.execute(sql % (
                    lecture.semester_date, lecture.code, lecture.name, lecture.grades, lecture.class_number,
                    lecture.department, lecture.professor, lecture.is_english))

                cur.execute("UPDATE koin.versions SET version = '%s_%d' WHERE type = 'timetable'" % (semester_date, int(time.time())))
                connection.commit()

    except Exception as error:
        connection.rollback()
        print(error)


class Lecture:
    def __init__(self, semester_date, code, name, grades, class_number, department, professor,
                 is_english):
        self.semester_date = semester_date
        self.code = code
        self.name = name
        self.grades = grades
        self.class_number = class_number
        self.department = department
        self.professor = professor
        self.is_english = is_english
        pass


if __name__ == "__main__":
    connection = connect_db()
    crawling()
    connection.close()
