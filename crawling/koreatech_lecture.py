import sys
import pymysql
import urllib3
import openpyxl
# from openpyxl.utils.cell import column_index_from_string
import config
import time
from datetime import date
from enum import Enum

OFFSET_BETWEEN_SCHEMA_INSTANCE = 1

SCHEMA_ROW_DELIMITER = 'No.'

sys.path.append("/home/ubuntu/myvenv/lib/python3.5/site-packages")


class ColumnNames(Enum):
    # 매핑이 되는 칼럼 명을 여러가지로 설정할 수 있도록 하였다.
    # `학\n점` 등으로 적혀져 있는 것을 주의할 것
    # 23.2 기준, 학교에서 "수강정원" 대신 "수정정원"으로 올려놓아서, 이에 대응하기 위해 배열로 여러 후보군을 설정하였다.
    # 25.1 기준, No. 행의 값이 없는 경우 생략하도록 추가
    ID = ["No."]
    SEMESTER = ["학기"]
    CODE = ["과목코드"]
    NAME = ["교과목명"]
    GRADES = ["학\n점", "학점"]
    CLASS_NUMBER = ["분반"]
    REGULAR_NUMBER = ["수강\n정원", "수정\n정원", "정원"]
    DEPARTMENT = ["개설학부(과)"]
    TARGET = ["수강신청\n가능학년", "수강대상 (학부/전공/학년)", "대상학부(과)"]
    PROFESSOR = ["담당교수"]
    IS_ENGLISH = ["영어강의"]
    DESIGN_SCORE = ["설\n계", "설계"]
    IS_ELEARNING = ["E-Learning"]
    CLASS_TIME = ["강의시간"]

    def include(self, column_name):
        return column_name in self.value


### static field ###
# 정규 수강신청 엑셀파일
year = date.today().year  # 오늘 연도
filename = 'lecture.xlsx'  # 읽어들일 엑셀파일명
day_to_index = {'월': '0', '화': '1', '수': '2', '목': '3', '금': '4', '토': '5', '일': '6'}


class WorkSheet:
    def __init__(self, work_sheet):
        self.sheet = work_sheet
        self.max_row = work_sheet.max_row
        self.instance_max_row = work_sheet.max_row
        self.max_column = work_sheet.max_column

        for i in range(1, self.max_row + 1):
            if self.at('A', i) == SCHEMA_ROW_DELIMITER:
                self.schema_row = i
                self.instance_start_row = i + OFFSET_BETWEEN_SCHEMA_INSTANCE
                break

        if self.instance_start_row == 0 or self.instance_start_row > self.instance_max_row:
            raise Exception(SCHEMA_ROW_DELIMITER + "를 통해 행 시작을 찾을 수 없습니다.")

    # A1 형식으로 셀에 접근 (col, row 순)
    def at(self, column, row):
        return self.sheet['%s%d' % (column, row)].value


class WorkSheetMapper:
    def __init__(self, work_sheet_helper):
        self.mapping_table = {}
        self.work_sheet_helper = work_sheet_helper

        row = work_sheet_helper.schema_row
        for i in range(1, work_sheet_helper.max_column + 1):
            col = self.get_column(i)

            self.mapping_for(col, row, work_sheet_helper)

        self.validates()

    def validates(self):
        if len(self.mapping_table) != len(ColumnNames):
            errors = ""

            for actual_column_name in self.mapping_table:
                for expected_column_name in ColumnNames:
                    if not expected_column_name.include(actual_column_name):
                        errors += "%s(%s) " % (expected_column_name.name, expected_column_name.value)

    def get_column(self, i):
        if i <= 26:
            return chr(i + 64)
        else:
            first_char = ((i - 1) // 26)
            second_char = ((i - 1) % 26) + 1
        return chr(first_char + 64) + chr(second_char + 64)


    def mapping_for(self, col, row, work_sheet_helper):
        for column_name in ColumnNames:
            if column_name.include(work_sheet_helper.at(col, row)):
                self.mapping_table[column_name] = col
                break

    def get(self, column_name, row_index):
        return self.work_sheet_helper.at(self.mapping_table[column_name], row_index)


def connect_db():
    urllib3.disable_warnings()
    conn = pymysql.connect(host=config.DATABASE_CONFIG['host'],
                           port=config.DATABASE_CONFIG['port'],
                           user=config.DATABASE_CONFIG['user'],
                           password=config.DATABASE_CONFIG['password'],
                           db=config.DATABASE_CONFIG['db'],
                           charset='utf8')
    return conn


def crawling():
    wb = openpyxl.load_workbook(filename=filename, data_only=True)
    ws = wb.active
    lectures = []
    work_sheet = WorkSheet(ws)
    work_sheet_mapper = WorkSheetMapper(work_sheet)

    semester = work_sheet_mapper.get(ColumnNames.SEMESTER, work_sheet.instance_start_row)
    semester_date = '%s%s' % (year, semester.split('학기')[0])

    for row in range(work_sheet.instance_start_row, work_sheet.max_row + 1):
        # No.에 해당하는 값이 없는 경우 건너 뛰기
        id = work_sheet_mapper.get(ColumnNames.ID, row)
        if not id:
            continue
            
        code = work_sheet_mapper.get(ColumnNames.CODE, row)
        name = work_sheet_mapper.get(ColumnNames.NAME, row)
        grades = work_sheet_mapper.get(ColumnNames.GRADES, row)
        class_number = work_sheet_mapper.get(ColumnNames.CLASS_NUMBER, row)
        regular_number = work_sheet_mapper.get(ColumnNames.REGULAR_NUMBER, row)
        if not regular_number:
            regular_number = ''
        department = work_sheet_mapper.get(ColumnNames.DEPARTMENT, row)
        if not department:
            department = ''
        target = work_sheet_mapper.get(ColumnNames.TARGET, row)
        if target:  # None이 아니라면 target의 여백들 지워준다.
            target = str(target).strip()
        else:
            target = ''
        professor = work_sheet_mapper.get(ColumnNames.PROFESSOR, row)
        if not professor:
            professor = ''
        is_english = work_sheet_mapper.get(ColumnNames.IS_ENGLISH, row)
        if not is_english:
            is_english = '0'
        design_score = work_sheet_mapper.get(ColumnNames.DESIGN_SCORE, row)
        is_elearning = work_sheet_mapper.get(ColumnNames.IS_ELEARNING, row)
        if not is_elearning:
            is_elearning = '0'
        class_time = convert_classtime(work_sheet_mapper.get(ColumnNames.CLASS_TIME, row))

        lecture = Lecture(semester_date=semester_date, code=code, name=name, grades=grades, class_number=class_number,
                          regular_number=regular_number, department=department, target=target,
                          professor=professor, is_english=is_english, design_score=design_score,
                          is_elearning=is_elearning, class_time=class_time)
        lectures.append(lecture)

        print(semester_date, code, name, grades, class_number, regular_number, department, target, professor,
              is_english, design_score, is_elearning, class_time)

    updateDB(lectures=lectures, semester_date=semester_date)
    pass


def convert_classtime(class_time):  # 강의 시간 변환 신버전
    classList = []
    day = ''
    try:
        for time in class_time.split(','):  # ,를 기준으로 파싱
            if time[0].isalpha():
                day = time[0]  # 요일
                time = time[1:]  # 요일을 제외한 강의 시간
            periodFlag = False
            period = time.split('~')  # ~를 기준으로 파싱
            timeTo = timeFrom = 0
            for p in period:
                result = "%s%02d" % (day_to_index.get(day, ''), 2 * (int(p[0:2]) - 1) + ord(p[2:3]) - ord('A'))  # 변환 공식
                if not periodFlag:
                    timeFrom = int(result)
                    periodFlag = True
                else:
                    timeTo = int(result)

            if not timeTo:  # timeTo가 할당되지 않은 예외 상황이라면
                timeTo = timeFrom  # 동일하게 할당

            for i in range(timeFrom, timeTo + 1):
                classList.append(i)
    except Exception:
        print('Error occured at: %s' % class_time)
        return []
    return classList


# def convert_classtime(ws, row):  # 강의 시간 변환 구버전

# start_index = column_index_from_string(class_time_col)
# class_time = []
# for column in range(start_index, start_index + 12):
#     detail_time = ws.cell(row=row, column=column).value
#     if not detail_time or not detail_time.strip():  # 빈 시간 제외
#         continue
#     day = detail_time.split('/')[0]
#     time = detail_time.split('/')[1]
#     if not day or not time:  # 누락된 시간 예외 처리
#         continue
#     try:
#         result = "%s%02d" % (day_to_index.get(day, ''), 2 * (int(time[0:2]) - 1) + ord(time[2:3]) - ord('A'))
#         class_time.append(int(result))
#     except Exception as error:
#         print(error)
# return class_time


def updateDB(lectures, semester_date):
    cur = connection.cursor()
    try:
        cur.execute("INSERT INTO koin.semester (SEMESTER) \
                        SELECT '%s' FROM DUAL \
                            WHERE NOT EXISTS (SELECT SEMESTER FROM koin.semester WHERE SEMESTER='%s')" % (
            semester_date, semester_date))

        cur.execute("DELETE FROM koin.lectures WHERE SEMESTER_DATE='%s'" % semester_date)

        for lecture in lectures:
            sql = "INSERT INTO koin.lectures(semester_date, code, name, grades, class, regular_number, department, target, professor, is_english, design_score, is_elearning, class_time) \
                        VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')"

            cur.execute(sql % (
                lecture.semester_date, lecture.code, lecture.name, lecture.grades, lecture.class_number,
                lecture.regular_number, lecture.department, lecture.target, lecture.professor,
                lecture.is_english, lecture.design_score, lecture.is_elearning, lecture.class_time))

        cur.execute(
            "UPDATE koin.versions SET version = '%s_%d' WHERE type = 'timetable'" % (semester_date, int(time.time())))
        connection.commit()

    except Exception as error:
        connection.rollback()
        print(error)


class Lecture:
    def __init__(self, semester_date, code, name, grades, class_number, regular_number, department, target, professor,
                 is_english, design_score, is_elearning, class_time):
        self.semester_date = semester_date
        self.code = code
        self.name = name
        self.grades = grades
        self.class_number = class_number
        self.regular_number = regular_number
        self.department = department
        self.target = target
        self.professor = professor
        self.is_english = is_english
        self.design_score = design_score
        self.is_elearning = is_elearning
        self.class_time = class_time
        pass


if __name__ == "__main__":
    try:
        connection = connect_db()
        crawling()
        connection.close()
    except Exception as error:
        print(error)
